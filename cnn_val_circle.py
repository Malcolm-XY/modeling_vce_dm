# -*- coding: utf-8 -*-
"""
Created on Sun Dec 22 01:23:16 2024

@author: 18307
"""
import os
import numpy as np
import pandas as pd

import torch

from utils import utils_feature_loading, utils_visualization
import cnn_validation

def cnn_cross_validation_circle(model, mapping_func, dataset, dist, resolution, interp, feature, subject_range, experiment_range):
    """"
    Currently only support SEED dataset.
    """
    labels = torch.tensor(np.array(utils_feature_loading.read_labels(dataset)))
    targets = labels.view(-1)
    distribution = utils_feature_loading.read_distribution(dataset, dist)
    
    results_entry = []
    for sub in subject_range:
        for ex in experiment_range:
            identifier = f'sub{sub}ex{ex}'
            print(f'Processing {identifier}...')

            cfs = utils_feature_loading.read_cfs(dataset, identifier, feature)
            cfs_alpha = cfs['alpha']
            cfs_beta = cfs['beta']
            cfs_gamma = cfs['gamma']


            alpha_mapped = mapping_func(cfs_alpha, distribution, resolution, interpolation=interp)
            beta_mapped = mapping_func(cfs_beta, distribution, resolution, interpolation=interp)
            gamma_mapped = mapping_func(cfs_gamma, distribution, resolution, interpolation=interp, imshow=True)
            
            data_mapped = np.stack((alpha_mapped, beta_mapped, gamma_mapped), axis=1)
            
            # Training and Validation
            # Run validation in no_grad mode to avoid unnecessary gradient computation
            result = cnn_validation.cnn_cross_validation(model, data_mapped, targets)
            
            # Add identifier to the result
            result['Identifier'] = f'sub{sub}ex{ex}'
            results_entry.append(result)
            
    # print(f'Final Results: {results_entry}')
    print('K-Fold Validation compelete\n')
    
    return results_entry

from openpyxl import load_workbook

def save_results_to_xlsx_append(results, output_dir, filename, sheet_name='K-Fold Results'):
    """
    Appends results to an existing Excel file or creates a new file if it doesn't exist.

    Args:
        results (list or pd.DataFrame): The results data to save.
        output_dir (str): The directory where the Excel file will be saved.
        filename (str): The name of the Excel file.
        sheet_name (str): The sheet name in the Excel file. Default is 'K-Fold Results'.

    Returns:
        str: The path of the saved Excel file.
    """
    # Convert results to DataFrame if necessary
    if not isinstance(results, pd.DataFrame):
        results_df = pd.DataFrame(results)
    else:
        results_df = results

    # Rearrange columns if "Identifier" is present
    if 'Identifier' in results_df.columns:
        columns_order = ['Identifier'] + [col for col in results_df.columns if col != 'Identifier']
        results_df = results_df[columns_order]

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Define the full output path
    output_path = os.path.join(output_dir, filename)

    # Append to existing Excel file or create a new one
    if os.path.exists(output_path):
        print(f"Appending data to existing file: {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the existing workbook
            existing_workbook = load_workbook(output_path)

            # Check if the sheet exists
            if sheet_name in existing_workbook.sheetnames:
                # Load existing sheet and append
                start_row = existing_workbook[sheet_name].max_row
                results_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
            else:
                # Write new sheet if not exists
                results_df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        print(f"Creating new file: {output_path}")
        results_df.to_excel(output_path, index=False, sheet_name=sheet_name)

    print(f"Results successfully saved to: {output_path}")
    return output_path

import time
import threading

def shutdown_with_countdown(countdown_seconds=30):
    """
    Initiates a shutdown countdown, allowing the user to cancel shutdown within the given time.

    Args:
        countdown_seconds (int): The number of seconds to wait before shutting down.
    """
    def cancel_shutdown():
        nonlocal shutdown_flag
        user_input = input("\nPress 'c' and Enter to cancel shutdown: ").strip().lower()
        if user_input == 'c':
            shutdown_flag = False
            print("Shutdown cancelled.")

    # Flag to determine whether to proceed with shutdown
    shutdown_flag = True

    # Start a thread to listen for user input
    input_thread = threading.Thread(target=cancel_shutdown, daemon=True)
    input_thread.start()

    # Countdown timer
    print(f"Shutdown scheduled in {countdown_seconds} seconds. Press 'c' to cancel.")
    for i in range(countdown_seconds, 0, -1):
        print(f"Time remaining: {i} seconds", end="\r")
        time.sleep(1)

    # Check the flag after countdown
    if shutdown_flag:
        print("\nShutdown proceeding...")
        os.system("shutdown /s /t 1")  # Execute shutdown command
    else:
        print("\nShutdown aborted.")

def end_program_actions(play_sound=True, shutdown=False, countdown_seconds=120):
    """
    Performs actions at the end of the program, such as playing a sound or shutting down the system.

    Args:
        play_sound (bool): If True, plays a notification sound.
        shutdown (bool): If True, initiates shutdown with a countdown.
        countdown_seconds (int): Countdown time for shutdown confirmation.
    """
    if play_sound:
        try:
            import winsound
            print("Playing notification sound...")
            winsound.Beep(1000, 500)  # Frequency: 1000Hz, Duration: 500ms
        except ImportError:
            print("winsound module not available. Skipping sound playback.")

    if shutdown:
        shutdown_with_countdown(countdown_seconds)

# %% read parameters
def read_params(model='exponential', model_fm='basic', model_rcm='differ'):
    token_model_fm = f'{model_fm}_models'
    if model_rcm == 'differ':
        token_rcm = 'DRCM'
    elif model_rcm == 'linear':
        token_rcm = 'LRCM'
    elif model_rcm == 'linear_ratio':
        token_rcm = 'LRRCM'
        
    path_current = os.getcwd()
    path_fitting_results = os.path.join(path_current, 'fitting_results')
    file_path = os.path.join(path_fitting_results, f'fitting_results({token_model_fm}, {token_rcm}, target_LDMI).xlsx')
    
    df = pd.read_excel(file_path).set_index('method')
    df_dict = df.to_dict(orient='index')
    
    model = model.upper()
    params = df_dict[model]
    
    return params

# %% Usage
import feature_engineering
import vce_modeling
from connectivity_matrix_rebuilding import cm_rebuilding as cm_rebuild
from models import models
def usage(feature_cm, model, model_fm, model_rcm, baseline=False):
    cnn_model = models.CNN_2layers_adaptive_maxpool_3()

    # labels
    labels = utils_feature_loading.read_labels(dataset='seed')
    y = torch.tensor(np.array(labels)).view(-1)
    
    # distance matrix
    _, dm = feature_engineering.compute_distance_matrix(dataset="seed", projection_params={"type": "3d"})
    utils_visualization.draw_projection(dm)

    param = read_params(model, model_fm, model_rcm)
    
    all_results_original = []
    all_results_rebuilded = []
    
    for sub in range(1, 16):
        for ex in range(1, 4):
            subject_id = f"sub{sub}ex{ex}"
            print(f"Evaluating {subject_id}...")
            
            # CM
            features = utils_feature_loading.read_fcs_mat('seed', subject_id, feature_cm)
            alpha = features['alpha']
            beta = features['beta']
            gamma = features['gamma']
            
            x = np.stack((alpha, beta, gamma), axis=1)

            # RCM
            alpha_rebuilded = cm_rebuild(alpha, dm, param, model, model_fm, model_rcm)
            beta_rebuilded = cm_rebuild(beta, dm, param, model, model_fm, model_rcm)
            gamma_rebuilded = cm_rebuild(gamma, dm, param, model, model_fm, model_rcm)
                
            x_rebuilded = np.stack((alpha_rebuilded, beta_rebuilded, gamma_rebuilded), axis=1)
            
            model_rebuilded = models.CNN_2layers_adaptive_maxpool_3()
            result_RCM = cnn_validation.cnn_cross_validation(model_rebuilded, x_rebuilded, y)
            
            # Add identifier to the result
            result_RCM['Identifier'] = f'sub{sub}ex{ex}'
            all_results_rebuilded.append(result_RCM)
            
            # baseline
            if baseline:
                model_original = models.CNN_2layers_adaptive_maxpool_3()
                result_CM = cnn_validation.cnn_cross_validation(model_original, x, y)
                # Add identifier to the result
                result_CM['Identifier'] = f'sub{sub}ex{ex}'
                all_results_original.append(result_CM)
            
    # print(f'Final Results: {results_entry}')
    print('K-Fold Validation compelete\n')
    
    # %% Save results to XLSX (append mode)
    output_dir = os.path.join(os.getcwd(), 'Results')
    filename_CM = f"cnn_validation_CM_{feature_cm.upper()}.xlsx"
    
    if model_fm == 'basic':
        token_model_fm = 'BFM'
    elif model_fm == 'advanced':
        token_model_fm = 'AFM'
    
    if model_rcm == 'differ':
        token_model_rcm = 'DRCM'
    elif model_rcm == 'linear':
        token_model_rcm = 'LRCM'
    elif model_rcm == 'linear_ratio':
        token_model_rcm = 'LRRCM'
    
    filename_RCM = f"cnn_validation_RCM_{feature_cm.upper()};{token_model_fm}({model})_{token_model_rcm}.xlsx"
    save_results_to_xlsx_append(all_results_original, output_dir, filename_CM)
    save_results_to_xlsx_append(all_results_rebuilded, output_dir, filename_RCM)
    
    return all_results_original, all_results_rebuilded

if __name__ == '__main__':
    model, model_fm, model_rcm = 'generalized_gaussian', 'advanced', 'linear'
    # params = read_params(model, model_fm, model_rcm)
    
    results_cm, results_rcm = usage('pcc', model, model_fm, model_rcm, baseline=False)
    results_cm_plv, results_rcm_plv = usage('plv', model, model_fm, model_rcm, baseline=False)
    
    # %% End
    end_program_actions(play_sound=True, shutdown=True, countdown_seconds=120)